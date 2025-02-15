from django.http import JsonResponse, HttpResponse
from django.conf import settings

import os
import openpyxl
import logging

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
# 핸들러 설정
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

# 포매터 설정
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)

# 핸들러 추가
logger.addHandler(console_handler)

def send_excel_content(request):

    # Excel 파일 경로 설정 (예: 프로젝트 루트 디렉토리의 'data.xlsx' 파일)
    global_vars = settings.GLOBAL_VARIABLE
    yes24_file_path = os.path.join(os.path.dirname(__file__), global_vars["yes24ExcelPath"])
    kobo_file_path = os.path.join(os.path.dirname(__file__), global_vars["koboExcelPath"])
    ypbooks_file_path = os.path.join(os.path.dirname(__file__), global_vars["ypbooksExcelPath"])
    aladin_file_path = os.path.join(os.path.dirname(__file__), global_vars["aladinExcelPath"])

    # 파일이 존재하는지 확인
    if os.path.exists(yes24_file_path):
        # Excel 파일 열기
        workbook = openpyxl.load_workbook(yes24_file_path)
        sheet = workbook.active

        # Excel 내용 읽기 (예: A1부터 C10까지의 데이터)
        json_object = {}
        excelDataList = []
        for row in sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=5, values_only=True):
            excelDataList.append(list(row))

        json_object['yes24_data'] = excelDataList

        workbook.close()

        excelDataList = []
        workbook = openpyxl.load_workbook(kobo_file_path)
        sheet = workbook.active

        # Excel 내용 읽기 (예: A1부터 C10까지의 데이터)
        for row in sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=5, values_only=True):
            excelDataList.append(list(row))

        json_object['kobo_data'] = excelDataList

        workbook.close()

        excelDataList = []
        workbook = openpyxl.load_workbook(ypbooks_file_path)
        sheet = workbook.active

        # Excel 내용 읽기 (예: A1부터 C10까지의 데이터)
        for row in sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=5, values_only=True):
            excelDataList.append(list(row))

        json_object['ybbooks_data'] = excelDataList

        # Excel 데이터를 JSON 응답으로 반환
        workbook.close()

        excelDataList = []
        workbook = openpyxl.load_workbook(aladin_file_path)
        sheet = workbook.active

        # Excel 내용 읽기 (예: A1부터 C10까지의 데이터)
        for row in sheet.iter_rows(min_row=2, max_row=6, min_col=1, max_col=5, values_only=True):
            excelDataList.append(list(row))

        json_object['aladin_data'] = excelDataList

        # Excel 데이터를 JSON 응답으로 반환
        logger.info('json_object')
        logger.info(json_object)

        workbook.close()

        return JsonResponse(json_object, safe=False)
    else:
        return HttpResponse("Crawling File not found", status=404)